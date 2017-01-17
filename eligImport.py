## imports ########################################################################
import csv
import itertools
import os
import glob
import pypyodbc as pyodbc
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import re
import sys
import time
from contextlib import contextmanager
import shutil
###################################################################################

## use the below global vars to set the file location and move location############
dest        = r"C:\Users\JMyers\Documents\humana data mart\imported\ "
dest        = dest + time.strftime("%Y%m%d")
file_loc    = r"C:\Users\JMyers\Documents\humana data mart\*.xlsx"
conn_string = r'DRIVER={SQL Server};SERVER=IDC-VMSQL-DEV02.HCP.COM;DATABASE=Catalyst_FullWell;Trusted_Connection=yes'
###################################################################################

## create helper functions ########################################################
def left(s, amount):
    if s is None:
        return ''
    else:
        return s[:amount]

def right(s, amount):
    if s is None:
        return ''
    else:
        return s[-amount:]

def xstr(s): ## fix blank data for SQL insert
    if s is None:
        return ''
    elif s == 'None':
        return ''
    else:
        return str(s)

def match(pattern, s): ## check s against pattern, return true if match, false otherwise
    if re.search(pattern, s):
        return True
    else:
        return False

def get_source(s):
    check_source = ''
    r_elig = "REMBX"
    r_clm  = "RECLM"
    r_ros  = "Roster"
    if match(r_elig, s):
        check_source = 'ELIG'
    elif match(r_clm, s):
        check_source = 'CLM'
    elif match(r_ros, s):
        check_source = 'PROV'
    else:
        check_source = 'None'
    return check_source

def clear_staging(s):
    ## need to re-write this so it only clears before a table is updated
    ## in case of asynch updates
    with open_db_connection(commit=True) as cursor:
        if s == 'ELIG':
            del_sql = "DELETE FROM dm_elig_staging"
        elif s == 'CLM':
            del_sql = "DELETE FROM dm_claims_staging"
        elif s == 'PROV':
            del_sql = "DELETE FROM dm_roster_staging"
        else:
            del_sql = ""
        cursor.execute(del_sql)
    print(s + ' cleared.')

def move_file(s):
    global dest
    if not os.path.exists(dest):
        os.makedirs(dest)
    origin = s
    try:
        shutil.move(origin, dest)
    except shutil.Error as err:
        error, = err.args
        sys.stderr.write(error.message)
        raise err
    print("\nFile Moved.")

@contextmanager
def open_db_connection(commit=False):
    global conn_string
    conn = pyodbc.connect(conn_string)
    cursor = conn.cursor()
    try:
        yield cursor
    except pyodbc.DatabaseError as err:
        error, = err.args
        sys.stderr.write(error.message)
        cursor.execute("ROLLBACK")
        raise err
    else:
        if commit:
            cursor.execute("COMMIT")
        else:
            cursor.execute("ROLLBACK")
    finally:
        conn.close()

def import_elig(s):
    clear_staging('ELIG')
    wb = load_workbook(s)
    ws = wb.active
    total_rows = ws.max_row
    rowcount = 2
    sql = ''
    with open_db_connection(commit=True) as cursor:
        for row in ws.rows:
            monthcode           = str(ws.cell(row = rowcount, column = 1).value).replace("'", "").strip()   ## rpt_pe
            lob_class           = str(ws.cell(row = rowcount, column = 2).value).replace("'", "").strip()   ## lob_class
            provider_name       = str(ws.cell(row = rowcount, column = 7).value).replace("'", "").strip()   ## center_nm
            provider_npi        = str(ws.cell(row = rowcount, column = 65).value).replace("'", "").strip()  ## prov_npi
            provider_contract   = str(ws.cell(row = rowcount, column = 6).value).replace("'", "").strip()   ## gk_ctrct
            member_name         = str(ws.cell(row = rowcount, column = 9).value).replace("'", "").strip()   ## mbrname
            member_id           = str(ws.cell(row = rowcount, column = 12).value).replace("'", "").strip()  ## mbr_id
            member_hpid         = str(ws.cell(row = rowcount, column = 13).value).replace("'", "").strip()  ## mbr_pid
            member_address      = str(ws.cell(row = rowcount, column = 15).value).replace("'", "").strip()  ## mb_addr2
            member_address+= ' '+ str(ws.cell(row = rowcount, column = 14).value).replace("'","").strip()   ## mbr_addr1
            member_city         = str(ws.cell(row = rowcount, column = 16).value).replace("'", "").strip()  ## mbr_city
            member_state        = str(ws.cell(row = rowcount, column = 17).value).replace("'", "").strip()  ## mbr_state
            member_zip          = str(ws.cell(row = rowcount, column = 18).value).replace("'", "").strip()  ## mbr_zip
            member_county       = str(ws.cell(row = rowcount, column = 20).value).replace("'", "").strip()  ## mbr_cntynm
            member_phone        = str(ws.cell(row = rowcount, column = 21).value).replace("'", "").strip()  ## new_phone2
            member_dob          = str(ws.cell(row = rowcount, column = 22).value).replace("'", "").strip()  ## mbr_bth
            member_dod          = str(ws.cell(row = rowcount, column = 64).value).replace("'", "").strip()  ## deceased_dt
            member_sex          = str(ws.cell(row = rowcount, column = 24).value).replace("'", "").strip()  ## mbr_sex
            member_enroll_date  = str(ws.cell(row = rowcount, column = 57).value).replace("'", "").strip()  ## product_eff_dt
            healthplan          = "H"                                                                       ## auto-assigned for now, will revisit with secure
            pro_lob             = str(ws.cell(row = rowcount, column = 47).value).replace("'","").strip()   ## pro_lob
            gk_adj_cd           = str(ws.cell(row = rowcount, column = 31).value).replace("'","").strip()   ## gk_adj_cd

            rowcount = rowcount + 1

            sql = """INSERT INTO dm_elig_staging 
                (monthcode, lob_class, provider_name, provider_npi, provider_contract, member_name, member_id, member_hpid, member_address,
                member_city, member_state, member_zip, member_county, member_phone, member_dob, member_dod, member_sex, member_enroll_date, healthplan,
                pro_lob, gk_adj_cd) 
                VALUES ('""" \
                + xstr(monthcode) + "','" \
                + xstr(lob_class) + "','" \
                + xstr(provider_name) + "','" \
                + xstr(provider_npi) + "','" \
                + xstr(provider_contract) + "','" \
                + xstr(member_name) + "','" \
                + xstr(member_id) + "','" \
                + xstr(member_hpid) + "','" \
                + xstr(member_address) + "','" \
                + xstr(member_city) + "'\n,'" \
                + xstr(member_state) + "','" \
                + xstr(member_zip) + "','" \
                + xstr(member_county) + "','" \
                + xstr(member_phone) + "','" \
                + xstr(member_dob) + "','" \
                + xstr(member_dod) + "','" \
                + xstr(member_sex) + "','" \
                + xstr(member_enroll_date) + "','" \
                + xstr(healthplan) + "','" \
                + xstr(pro_lob) + "','" \
                + xstr(gk_adj_cd) +  "')"
            ##print(sql) ## uncomment if insert fails with an error to identify fail line
            print(str(rowcount) + ' rows loaded out of ' + str(total_rows) + ' total rows.', end='\r')
            cursor.execute(sql)
    move_file(s)
    print("\n")

def import_roster(s):
    clear_staging('PROV')
    wb = load_workbook(s)
    ws = wb.active
    total_rows = ws.max_row
    rowcount = 2
    sql = ''
    with open_db_connection(commit=True) as cursor:
        for row in ws.rows:
            provider_npi     = str(ws.cell(row = rowcount, column = 1).value).replace("'", "").strip()   ## provider npi
            region           = str(ws.cell(row = rowcount, column = 8).value).replace("'", "").strip()   ## region
            contract_entity  = str(ws.cell(row = rowcount, column = 7).value).replace("'", "").strip()   ## contract / physician entity
            eff_date         = str(ws.cell(row = rowcount, column = 9).value).replace("'", "").strip()   ## effective date
            type_r           = str(ws.cell(row = rowcount, column = 10).value).replace("'", "").strip()  ## pcp/specialist
            specialty        = str(ws.cell(row = rowcount, column = 11).value).replace("'", "").strip()  ## primary specialty
            practice_name    = str(ws.cell(row = rowcount, column = 13).value).replace("'", "").strip()  ## primary practice name
            practice_npi     = str(ws.cell(row = rowcount, column = 14).value).replace("'", "").strip()  ## practice npi
            practice_address = str(ws.cell(row = rowcount, column = 15).value).replace("'", "").strip()  ## primary address
            practice_city    = str(ws.cell(row = rowcount, column = 17).value).replace("'","").strip()   ## primary city
            practice_state   = str(ws.cell(row = rowcount, column = 18).value).replace("'", "").strip()  ## primary state
            practice_zip     = str(ws.cell(row = rowcount, column = 19).value).replace("'", "").strip()  ## primary zip
            practice_county  = str(ws.cell(row = rowcount, column = 20).value).replace("'", "").strip()  ## primary county
            practice_phone   = str(ws.cell(row = rowcount, column = 21).value).replace("'", "").strip()  ## primary phone
            practice_tin     = str(ws.cell(row = rowcount, column = 23).value).replace("'", "").strip()  ## practice tin
            network          = ""  ## none

            rowcount = rowcount + 1

            check = u"\u2019"
            if re.search(practice_npi, check):
                practice_npi = -1
            elif isinstance(right(practice_npi, 1), (str, bytes)): ## account for letter suffixes added to CHN roster NPIs
                practice_npi = left(practice_npi, 10)

            sql = """INSERT INTO dm_roster_staging 
                (provider_npi, region, contract_entity, eff_date, type, specialty, practice_name, practice_npi, practice_address
                , practice_city, practice_state, practice_zip, practice_county, practice_phone, practice_tin, network) 
                VALUES ('""" \
                + xstr(provider_npi) + "','" \
                + xstr(region) + "','" \
                + xstr(contract_entity) + "','" \
                + xstr(eff_date) + "','" \
                + xstr(type_r) + "','" \
                + xstr(specialty) + "','" \
                + xstr(practice_name) + "','" \
                + xstr(practice_npi) + "','" \
                + xstr(practice_address) + "','" \
                + xstr(practice_city) + "'\n,'" \
                + xstr(practice_state) + "','" \
                + xstr(practice_zip) + "','" \
                + xstr(practice_county) + "','" \
                + xstr(practice_phone) + "','" \
                + xstr(practice_tin) + "','" \
                + xstr(network) + "')"
            ##print(sql) ## uncomment if insert fails with an error to identify fail line
            print(str(rowcount) + ' rows loaded out of ' + str(total_rows) + ' total rows.', end='\r')
            cursor.execute(sql)
    move_file(s)
    print("\n")

def import_claims(s):
    clear_staging('CLM')
    wb = load_workbook(s) ## open elig workbook
    ws = wb.active
    total_rows = ws.max_row
    rowcount = 2
    sql = ''
    with open_db_connection(commit=True) as cursor:
        for row in ws.rows:
            reporting_period    = str(ws.cell(row = rowcount, column = 148).value).replace("'", "").strip()   ## pe_date
            member_pid          = str(ws.cell(row = rowcount, column = 146).value).replace("'", "").strip()   ## mbr_pid
            member_enroll_id    = str(ws.cell(row = rowcount, column = 15).value).replace("'", "").strip()    ## sub_id
            patient_name        = str(ws.cell(row = rowcount, column = 19).value).replace("'", "").strip()    ## pat_name
            patient_dob         = str(ws.cell(row = rowcount, column = 21).value).replace("'", "").strip()    ## pat_dob
            primary_prov_id     = str(ws.cell(row = rowcount, column = 43).value).replace("'", "").strip()    ## prim_prov
            rend_prov_id        = str(ws.cell(row = rowcount, column = 44).value).replace("'", "").strip()    ## serv_prov
            rend_prov_name      = str(ws.cell(row = rowcount, column = 45).value).replace("'", "").strip()    ## sprov_nm
            rend_prov_zip       = str(ws.cell(row = rowcount, column = 53).value).replace("'", "").strip()    ## sprov_zip
            gk_provider_name    = str(ws.cell(row = rowcount, column = 141).value).replace("'", "").strip()   ## gk_prov_nm
            gk_prov_contract    = str(ws.cell(row = rowcount, column = 140).value).replace("'", "").strip()   ## gk_prov_id
            refer_prov_id       = str(ws.cell(row = rowcount, column = 54).value).replace("'", "").strip()    ## refer_prov
            claim_no            = str(ws.cell(row = rowcount, column = 10).value).replace("'", "").strip()    ## claim_nbr
            claim_type          = str(ws.cell(row = rowcount, column = 11).value).replace("'", "").strip()    ## claim_type
            rend_prov_type      = str(ws.cell(row = rowcount, column = 145).value).replace("'", "").strip()   ## prov_type
            admit_source        = str(ws.cell(row = rowcount, column = 24).value).replace("'", "").strip()    ## adm_source
            admit_date          = str(ws.cell(row = rowcount, column = 25).value).replace("'", "").strip()    ## adm_date
            discharge_date      = str(ws.cell(row = rowcount, column = 29).value).replace("'","").strip()     ## dischrg_dt
            admit_dx            = str(ws.cell(row = rowcount, column = 28).value).replace("'","").strip()     ## adm_dx
            discharge_status    = str(ws.cell(row = rowcount, column = 30).value).replace("'","").strip()     ## dischrg_st
            ndc_code            = str(ws.cell(row = rowcount, column = 31).value).replace("'","").strip()     ## ndc_nbr
            rx_quantity         = str(ws.cell(row = rowcount, column = 32).value).replace("'","").strip()     ## rx_qty
            service_from_date   = str(ws.cell(row = rowcount, column = 88).value).replace("'","").strip()     ## srv_frm_dt
            service_to_date     = str(ws.cell(row = rowcount, column = 89).value).replace("'","").strip()     ## srv_to_dt
            process_date        = str(ws.cell(row = rowcount, column = 34).value).replace("'","").strip()     ## proc_dt
            service_type_code   = str(ws.cell(row = rowcount, column = 109).value).replace("'","").strip()    ## type
            procedure_desc      = str(ws.cell(row = rowcount, column = 183).value).replace("'","").strip()    ## proc_desc
            service_units       = str(ws.cell(row = rowcount, column = 91).value).replace("'","").strip()     ## serv_units
            dx_code_1           = str(ws.cell(row = rowcount, column = 58).value).replace("'","").strip()     ## diag_code_1
            dx_code_2           = str(ws.cell(row = rowcount, column = 59).value).replace("'","").strip()     ## diag_code_2
            dx_code_3           = str(ws.cell(row = rowcount, column = 60).value).replace("'","").strip()     ## diag_code_3
            dx_code_4           = str(ws.cell(row = rowcount, column = 61).value).replace("'","").strip()     ## diag_code_4
            dx_code_5           = str(ws.cell(row = rowcount, column = 62).value).replace("'","").strip()     ## diag_code_5
            dx_code_6           = str(ws.cell(row = rowcount, column = 63).value).replace("'","").strip()     ## diag_code_6
            dx_code_7           = str(ws.cell(row = rowcount, column = 64).value).replace("'","").strip()     ## diag_code_7
            dx_code_8           = str(ws.cell(row = rowcount, column = 65).value).replace("'","").strip()     ## diag_code_8
            dx_code_9           = str(ws.cell(row = rowcount, column = 66).value).replace("'","").strip()     ## diag_code_9
            proc_code_1         = str(ws.cell(row = rowcount, column = 76).value).replace("'","").strip()     ## proc_code_1
            proc_code_2         = str(ws.cell(row = rowcount, column = 77).value).replace("'","").strip()     ## proc_code_2
            proc_code_3         = str(ws.cell(row = rowcount, column = 78).value).replace("'","").strip()     ## proc_code_3
            proc_code_4         = str(ws.cell(row = rowcount, column = 79).value).replace("'","").strip()     ## proc_code_4
            proc_code_5         = str(ws.cell(row = rowcount, column = 80).value).replace("'","").strip()     ## proc_code_5
            proc_code_6         = str(ws.cell(row = rowcount, column = 81).value).replace("'","").strip()     ## proc_code_6
            drg_no              = str(ws.cell(row = rowcount, column = 90).value).replace("'","").strip()     ## drg_nbr
            pot_code            = str(ws.cell(row = rowcount, column = 92).value).replace("'","").strip()     ## pot
            claim_charge        = str(ws.cell(row = rowcount, column = 93).value).replace("'","").strip()     ## chrg_amt
            claim_paid          = str(ws.cell(row = rowcount, column = 96).value).replace("'","").strip()     ## act_pd_amt
            claim_discount      = str(ws.cell(row = rowcount, column = 102).value).replace("'","").strip()    ## discnt_amt
            lob_code            = str(ws.cell(row = rowcount, column = 128).value).replace("'","").strip()    ## plan_lob
            generic_ind         = str(ws.cell(row = rowcount, column = 154).value).replace("'","").strip()    ## gen_ind
            pharm_npi           = str(ws.cell(row = rowcount, column = 161).value).replace("'","").strip()    ## phar_npi
            pharm_name          = str(ws.cell(row = rowcount, column = 162).value).replace("'","").strip()    ## phar_nm
            
            if not isinstance(pharm_npi, int): ## account for "PAPERCLAIM" item in pharmacy npi values
                pharm_npi = -1

            rowcount = rowcount + 1
            
            sql += """INSERT INTO dm_claims_staging 
                (reporting_period, member_pid, member_enroll_id, patient_name, patient_dob, primary_prov_id, rend_prov_id
                , rend_prov_name, rend_prov_zip, gk_provider_name, gk_prov_contract, refer_prov_id, claim_no, claim_type
                , rend_prov_type, admit_source, admit_date, discharge_date, admit_dx, discharge_status, ndc_code, rx_quantity
                , service_from_date, service_to_date, process_date, service_type_code, procedure_desc, service_units
                , dx_code_1, dx_code_2, dx_code_3, dx_code_4, dx_code_5, dx_code_6, dx_code_7, dx_code_8, dx_code_9
                , proc_code_1, proc_code_2, proc_code_3, proc_code_4, proc_code_5, proc_code_6, drg_no, pot_code, claim_charge
                , claim_paid, claim_discount, lob_code, generic_ind, pharm_npi, pharm_name) 
                VALUES ('""" \
                    + xstr(reporting_period) + "','" \
                    + xstr(member_pid) + "','" \
                    + xstr(member_enroll_id) + "','" \
                    + xstr(patient_name) + "','" \
                    + xstr(patient_dob) + "','" \
                    + xstr(primary_prov_id) + "','" \
                    + xstr(rend_prov_id) + "','" \
                    + xstr(rend_prov_name) + "','" \
                    + xstr(rend_prov_zip) + "','" \
                    + xstr(gk_provider_name) + "'\n,'" \
                    + xstr(gk_prov_contract) + "','" \
                    + xstr(refer_prov_id) + "','" \
                    + xstr(claim_no) + "','" \
                    + xstr(claim_type) + "','" \
                    + xstr(rend_prov_type) + "','" \
                    + xstr(admit_source) + "','" \
                    + xstr(admit_date) + "','" \
                    + xstr(discharge_date) + "','" \
                    + xstr(admit_dx) + "','" \
                    + xstr(discharge_status) + "','" \
                    + xstr(ndc_code) + "','" \
                    + xstr(rx_quantity) + "','" \
                    + xstr(service_from_date) + "','" \
                    + xstr(service_to_date) + "','" \
                    + xstr(process_date) + "','" \
                    + xstr(service_type_code) + "','" \
                    + xstr(procedure_desc) + "','" \
                    + xstr(service_units) + "'\n,'" \
                    + xstr(dx_code_1) + "','" \
                    + xstr(dx_code_2) + "','" \
                    + xstr(dx_code_3) + "','" \
                    + xstr(dx_code_4) + "','" \
                    + xstr(dx_code_5) + "','" \
                    + xstr(dx_code_6) + "','" \
                    + xstr(dx_code_7) + "','" \
                    + xstr(dx_code_8) + "'\n,'" \
                    + xstr(dx_code_9) + "','" \
                    + xstr(proc_code_1) + "','" \
                    + xstr(proc_code_2) + "','" \
                    + xstr(proc_code_3) + "','" \
                    + xstr(proc_code_4) + "','" \
                    + xstr(proc_code_5) + "','" \
                    + xstr(proc_code_6) + "','" \
                    + xstr(drg_no) + "'\n,'" \
                    + xstr(pot_code) + "','" \
                    + xstr(claim_charge) + "','" \
                    + xstr(claim_paid) + "','" \
                    + xstr(claim_discount) + "','" \
                    + xstr(lob_code) + "','" \
                    + xstr(generic_ind) + "','" \
                    + xstr(pharm_npi) + "','" \
                    + xstr(pharm_name) +  "');"
            ##print(sql) ## uncomment if insert fails with an error to identify fail line
            print(str(rowcount) + ' rows loaded out of ' + str(total_rows) + ' total rows.', end='\r')
            if rowcount % 100 == 0:
                cursor.execute(sql)
    move_file(s)
    print("\n")
###################################################################################

## enter main #####################################################################

file_list = glob.glob(file_loc) # load all xlsx files in file_loc to file_list

## begin iteration over all files
file_len = len(file_list)

if file_len == 0:
    print("No new files to import.")
else:
    print(str(file_len) + " new files to import.")
    
for x in range(0, file_len):
    print(os.path.basename(file_list[x]))
    check_source = get_source(os.path.basename(file_list[x]))
    print(check_source)
    if check_source == 'ELIG':
        import_elig(file_list[x])
    elif check_source == 'CLM':
        import_claims(file_list[x])
    elif check_source == 'PROV':
        import_roster(file_list[x])
    else:
        print("Unknown file: ?", file_list[x])
print("\nBeginning merge steps...") ## next step, add in stored proc runs and checks/validations



## config connection - delete once working
## conn = pyodbc.connect('DRIVER={SQL Server};SERVER=IDC-VMSQL-DEV02.HCP.COM;DATABASE=Catalyst_FullWell;Trusted_Connection=yes')
## $sa_Catalyst / Cs9KdL]syglav2vJL=\C
## cursor = conn.cursor()
